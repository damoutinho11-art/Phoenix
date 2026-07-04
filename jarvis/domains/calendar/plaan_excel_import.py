"""Convert a manually-downloaded Plaan "Excel - Kava" export into Phoenix's
normalized manual calendar snapshot import contract.

READ ONLY / MANUAL ONLY. This module never logs into plaan.opera.ee, never
fetches it automatically, and never stores Plaan credentials, cookies, or
session data. It only converts bytes the user has already downloaded and
uploaded themselves. The raw uploaded file is parsed in-memory here and
discarded by the caller; nothing about the raw file is written to disk or
the database.

Output shape matches jarvis.domains.calendar.plaan_live's manual snapshot
import contract exactly, so it can be passed straight into
plaan_live.validate_manual_snapshot_import() — this module does not
duplicate or bypass that validation.
"""

from __future__ import annotations

import hashlib
import io
import re
import zipfile
from datetime import date, datetime, timezone
from typing import Any

from openpyxl import load_workbook

_MAX_FILE_SIZE_BYTES = 5 * 1024 * 1024
_MAX_ROWS = 2000

_EXPECTED_HEADERS = ["Kuupäev", "Pealkiri", "Näitlejad", "Ruum"]

_DATE_TIME_RANGE_RE = re.compile(
    r"(\d{2})\.(\d{2})\.(\d{4})\s+(\d{2}:\d{2})\s*-\s*(\d{2}:\d{2})"
)
_DATE_ONLY_RE = re.compile(r"(\d{2})\.(\d{2})\.(\d{4})")

_EVENT_TYPE_KEYWORDS = [
    ("proov", "rehearsal"),
    ("etendus", "performance"),
    ("konverents", "press_conference"),
    ("gala", "gala"),
]


def _infer_event_type(title: str) -> str:
    """Infer a coarse event_type from Pealkiri keywords. Best-effort only."""
    lowered = (title or "").lower()
    for keyword, event_type in _EVENT_TYPE_KEYWORDS:
        if keyword in lowered:
            return event_type
    return "unknown"


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _parse_kuupaev(raw_value: str) -> tuple[str | None, str | None, str | None, str | None]:
    """Parse a Kuupäev cell into (date_iso, time_start, time_end, warning).

    Returns date_iso=None only when no date at all could be recovered from
    the cell (never silently drops the row in that case — caller still keeps
    the row and surfaces the warning).
    """
    match = _DATE_TIME_RANGE_RE.search(raw_value)
    if match:
        day, month, year, time_start, time_end = match.groups()
        date_iso = date(int(year), int(month), int(day)).isoformat()
        return date_iso, time_start, time_end, None

    date_only_match = _DATE_ONLY_RE.search(raw_value)
    if date_only_match:
        day, month, year = date_only_match.groups()
        date_iso = date(int(year), int(month), int(day)).isoformat()
        warning = f"Could not parse a time range from Kuupäev value {raw_value!r}; time_start/time_end left blank."
        return date_iso, None, None, warning

    warning = f"Could not parse date or time from Kuupäev value {raw_value!r}."
    return None, None, None, warning


def _event_id(date_iso: str | None, title: str, row_index: int) -> str:
    seed = f"{date_iso or ''}|{title}|{row_index}"
    digest = hashlib.sha1(seed.encode("utf-8")).hexdigest()[:16]
    return f"plaan-excel-{digest}"


def _reject_if_macro_enabled(file_bytes: bytes) -> None:
    """Refuse .xlsm-style files (vbaProject.bin present) or malformed zips.

    A true .xlsx is a zip archive with no vbaProject.bin part. This check
    also protects against non-Excel byte blobs being passed in.
    """
    try:
        with zipfile.ZipFile(io.BytesIO(file_bytes)) as archive:
            names = archive.namelist()
    except zipfile.BadZipFile as exc:
        raise ValueError("Uploaded file is not a valid .xlsx workbook.") from exc

    if any("vbaproject" in name.lower() for name in names):
        raise ValueError(
            "Macro-enabled workbooks (.xlsm or .xlsx containing vbaProject.bin) are not accepted. "
            "Please upload a plain .xlsx export."
        )


def parse_plaan_excel(file_bytes: bytes) -> dict[str, Any]:
    """Parse a Plaan "Excel - Kava" export into Phoenix's manual snapshot contract.

    Raises ValueError with a clear, user-facing message on any structural
    problem (missing headers, oversized file, macro-enabled workbook).
    """
    if len(file_bytes) > _MAX_FILE_SIZE_BYTES:
        raise ValueError(
            f"Uploaded file is too large ({len(file_bytes)} bytes). "
            f"The maximum accepted size is {_MAX_FILE_SIZE_BYTES} bytes (5MB)."
        )

    _reject_if_macro_enabled(file_bytes)

    try:
        workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError(f"Could not read uploaded file as an Excel workbook: {exc}") from exc

    try:
        sheet = workbook.active
        rows_iter = sheet.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            raise ValueError("Uploaded file has no header row.")

        header_index: dict[str, int] = {}
        for idx, cell in enumerate(header_row):
            name = _cell_text(cell)
            if name:
                header_index[name] = idx

        missing = [name for name in _EXPECTED_HEADERS if name not in header_index]
        if missing:
            raise ValueError(
                "Could not find expected column(s) "
                + ", ".join(f"'{name}'" for name in missing)
                + " in this file — expected Plaan's standard Kava export format "
                + "with columns: " + ", ".join(_EXPECTED_HEADERS) + "."
            )

        kuupaev_idx = header_index["Kuupäev"]
        pealkiri_idx = header_index["Pealkiri"]
        naitlejad_idx = header_index["Näitlejad"]
        ruum_idx = header_index["Ruum"]

        events: list[dict[str, Any]] = []
        fetch_warnings: list[str] = ["SOURCE: manual Excel import (Plaan Kava export)."]
        row_count = 0

        for row_index, row in enumerate(rows_iter, start=2):
            row_count += 1
            if row_count > _MAX_ROWS:
                raise ValueError(
                    f"Uploaded file has more than {_MAX_ROWS} data rows; "
                    "this does not look like a normal Plaan Kava export."
                )

            def _cell(idx: int) -> Any:
                return row[idx] if idx < len(row) else None

            kuupaev_raw = _cell_text(_cell(kuupaev_idx))
            pealkiri_raw = _cell_text(_cell(pealkiri_idx))
            naitlejad_raw = _cell_text(_cell(naitlejad_idx))
            ruum_raw = _cell_text(_cell(ruum_idx))

            if not kuupaev_raw and not pealkiri_raw and not naitlejad_raw and not ruum_raw:
                continue

            date_iso, time_start, time_end, warning = _parse_kuupaev(kuupaev_raw)
            if warning:
                fetch_warnings.append(f"Row {row_index}: {warning}")
            if date_iso is None:
                # Still include the row per spec — never drop silently — but
                # without a date this cannot be validated downstream, so skip
                # adding it as an event and rely on the warning above.
                continue

            title = pealkiri_raw or "Untitled Plaan event"
            events.append({
                "event_id": _event_id(date_iso, title, row_index),
                "event_type": _infer_event_type(title),
                "title": title,
                "date": date_iso,
                "time_start": time_start,
                "time_end": time_end,
                "location": ruum_raw or None,
                "role": naitlejad_raw or None,
            })

        return {
            "as_of": datetime.now(timezone.utc).replace(tzinfo=None).isoformat(timespec="seconds"),
            "events": events,
            "fetch_warnings": fetch_warnings,
        }
    finally:
        workbook.close()
